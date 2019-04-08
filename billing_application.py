from tkinter import *
from tkinter import font,messagebox
from PIL import ImageTk,Image
from escpos.connections import getUSBPrinter
import openpyxl

#printer = getUSBPrinter()(idVendor=0x1504,idProduct=0x0006,inputEndPoint=0x82,outputEndPoint=0x01)

book = openpyxl.load_workbook('item_details.xlsx')
sheet = book.active
work_sheet = book.get_sheet_by_name(name='Sheet1')
item_names=[]
for i in range(2,work_sheet.max_row+1):
    cell_obj=work_sheet.cell(row=i,column=2)
    item_names.append(cell_obj.value)


app=Tk()
app.minsize(1100,680)

app.configure(bg='lightgrey')
def_font=font.Font(family='georgia',size=13)
app.title('NAD Groups')
app.iconbitmap(r'6.ico')

def cancel(Event=None):
    global entry2,entry1,c1,list_box,qty
    list_box.delete(0,END)
    entry1.focus_set()
    search_name.set('')
    qty.set('')
    c1.delete('all')

def remove():
    global  bill_list,bill,entry1,bill_sno,total_amount
    
    if len(bill)==0:
        return
    try:
        i=bill_list.curselection()[0]
        bill_list.delete(0)
        bill_sno-=1
        total_amount-=bill[i][3]
        add_amount(0)
        bill.pop(i)
    except:
        messagebox.showinfo('Selection Required','First select the item!')
        return

    for j in range(i,len(bill)):
        bill[i][0]=bill[i][0]-1
    bill_list.delete(0,END)
    bill_list.insert(END,*bill)
    entry1.focus_set()

def clear():
    global bill,bill_sno,bill_list,paid,c2,entry2,entry1,c2,c3

    paid.set('')
    bill_list.delete(0,END)
    qty.set('')
    c1.delete('all')
    c2.delete('all')
    c3.delete('all')
    entry1.focus_set()
    bill=[]
    bill_sno=0
    bill_list.delete(0,END)

def hover(Event):
    global clear_btn,search_btn,remove_btn,add_btn,cancel_btn,bill_btn,stock_btn,update_btn
    h=Event.widget.cget('text')
    if h=='Clear':
        clear_btn['bg']='#939393'
        clear_btn['relief']='sunken'
    elif h=='Remove':
        remove_btn['bg']='#939393'
        remove_btn['relief']='sunken'
    elif h=='Cancel':
        cancel_btn['bg']='#939393'
        cancel_btn['relief']='sunken'
    elif h=='Search':
        search_btn['bg']='#939393'
        search_btn['relief']='sunken'
    elif h=='Add':
        add_btn['bg']='#939393'
        add_btn['relief']='sunken'
    elif h=='Print':
        print_btn['bg']='#939393'
        print_btn['relief']='sunken'
    elif h=='Bill':
        bill_btn['bg']='#7f95d8'
        bill_btn['relief']='sunken'
    elif h=='stock':
        stock_btn['bg']='#7f95d8'
        stock_btn['relief']='sunken'
    elif h=='Update':
        update_btn['bg']='#939393'
        update_btn['relief']='sunken'

def cancel_hover(Event):
    global clear_btn,search_btn,remove_btn,add_btn,cancel_btn,bill_btn,stock_btn,update_btn
    h=Event.widget.cget('text')
    if h=='Clear':
        clear_btn['bg']='#7f95d8'
        clear_btn['relief']='flat'
    elif h=='Remove':
        remove_btn['bg']='#7f95d8'
        remove_btn['relief']='flat'
    elif h=='Cancel':
        cancel_btn['bg']='#7f95d8'
        cancel_btn['relief']='flat'
    elif h=='Add':
        add_btn['bg']='#7f95d8'
        add_btn['relief']='flat'
    elif h=='Search':
        search_btn['bg']='#7f95d8'
        search_btn['relief']='flat'
    elif h=='Print':
        print_btn['bg']='#7f95d8'
        print_btn['relief']='flat'
    elif h=='Bill':
        bill_btn['bg']='#939393'
        bill_btn['relief']='flat'
    elif h=='stock':
        stock_btn['bg']='#939393'
        stock_btn['relief']='flat'
    elif h=='Update':
        update_btn['bg']='#7f95d8'
        update_btn['relief']='flat'

def search(s):
    global item_names
    temp=[]
    for i in range(len(item_names)):
        if s in item_names[i]:
            temp.append(item_names[i])
    return (temp)

def add_amount(val):
    global c2,total_amount
    total_amount+=val
    c2.delete('all')
    c2.create_text(33,12,text='Rs,'+str(total_amount))

def create_bill(i_name,q):
    global item_names,bill,bill_sno,bill_list,out,work_sheet,discount
    s1=''

    for i in range(1,len(item_names)):
        if item_names[i]==item_name:
            if q>work_sheet.cell(row=i+1,column=4).value:
                b='There are only '+str(int((work_sheet.cell(row=i+1,column=4)).value))+' avialable!'
                messagebox.showinfo('No stock',b)
                return 0
            else:
                bill_sno+=1
                d=(int(discount.get())*work_sheet.cell(row=i+1,column=3).value)/100
                bill.append([bill_sno,item_name,q,q*((work_sheet.cell(row=i+1,column=3).value)-d)])
                s1=(str(bill_sno)+'.').rjust(4)
                s2=item_name.ljust(10)
                s4=str(q).rjust(4)
                s3=str(q*((work_sheet.cell(row=i+1,column=3).value)-d)).rjust(8)
                s1=s1+s2+s4+s3+(work_sheet.cell(row=i+1,column=5).value).rjust(3)
                bill_list.insert(END,s1)
                bill_list.see(bill_sno)
                add_amount(q*((work_sheet.cell(row=i+1,column=3)).value))
                out+=s1[:-3]+'\n'
                discount.set(0)
                return 1


def get_quantity(Event=None):
    global entry2,item_name,entry1,search_name,qty

    q=int(qty.get())
    if q=='':
        return
    f=create_bill(item_name,q)
    if f==0:
        entry2.focus_set()
    else:
        entry1.focus_set()

def print_item(i_name):
    global c1,entry2,qty,up_cost

    qty.set('')
    c1.delete('all')
    c1.create_text(55,12,text='> '+i_name+'   Rs,'+str(work_sheet.cell(row=item_names.index(i_name)+1,column=3).value))
    up_cost.set(work_sheet.cell(row=item_names.index(i_name)+1,column=3).value)
    entry2.focus_set()

def get_item(Event=None):
    global list_box,entry2,item_name
    try:
        item_name=list_box.get(list_box.curselection())[3:]
        print_item(item_name)
    except:
        messagebox.showinfo('Selection Required','First select the item!')

def search_item(Event=None):
    global list_box,entry2,x1,list_box,l9,entry1
    l9.destroy()

    list_box.delete(0,END)
    x1.destroy()

    s_key=str(search_name.get()).upper()
    if s_key=='':
        return

    searched_items=search(s_key)
    if len(searched_items)==0:
        l9=Label(list_box,text='No such item found',font=('times new roman',11),bg='white')
        l9.place(relx=0.3,rely=0.4)
        entry1.focus_set()
        return

    for i in searched_items:
        list_box.insert(END,' > '+i)

    return

def get_balance(Event=None):
    global c3,entry3,paid,total_amount,entry1

    balance=int(paid.get())-total_amount
    c3.create_text(33,12,text='Rs,'+str(balance))
    entry1.focus_set()


def reduce_count(b):
    for j in range(len(b)):
        for i in range(2,work_sheet.max_row+1):
            if b[j][1]==work_sheet.cell(row=i,column=2).value:
                work_sheet.cell(row=i,column=4).value-=b[j][2]
                book.save('D:\python\project one\item_details.xlsx')
    return

def print_bill(b):
    pass

def get_money():
    global bill,bill_sno,bill_list,list_box,out,paid,total_amount,c3,entry3,total_amount,search_name,l8,income
    
    out+=('Total Amount : '+str(float(total_amount))).rjust(26)
    income.set(int(income.get())+total_amount)
    l8.destroy()
    l8=Label(app,text='Today`s total collection: Rs,'+str(income.get()),font=('times new roman',13),fg='white',bg='#4d4d4d')
    l8.place(relx=0.75,rely=0.038)
    entry3.focus_set()
    print_bill(bill)
    paid.set('')

    list_box.delete(0,END)
    print(bill)
    reduce_count(bill)
    print(out)
    bill=[]
    bill_sno=0
    search_name.set('')
    out=''

def item_search(*args):
    global list_box
    list_box.focus_set()

def callback(*args):
    search_item(None)

def billing():
    global work_sheet,income,discount,flag1,flag2,f6,l8,c7,logo,l1,l2,bill_btn,stock_btn,f0,f1,search_name,f5,lb1,lb2,lb4,lb3,lb5,lb6,lb7,entry1,entry2,entry3,c1,c2,c3,x1,list_box,bill_list,qty,clear_btn,remove_btn,cancel_btn,add_btn,search_btn,print_btn,paid

    if flag1==True:
        flag1=False
        flag2=True
    else:
        return

    search_name.set('')
    qty.set('')
    discount.set(0)
    f0=Frame(app,bg='lightgrey')
    f0.place(x=87,y=81,relheight=1,relwidth=1)
    logo=Label(f0,image=pic,bg='lightgrey')
    logo.place(x=-100,y=-100,relwidth=1,relheight=1)
    Label(f0,text='Billing',font=def_font,bg='#9fabcf').place(x=-50,relwidth=1)
    l8=Label(app,text='Today`s total collection: Rs,'+str(income.get()),font=('times new roman',13),fg='white',bg='#4d4d4d')
    l8.place(relx=0.75,rely=0.038)
    f1=Frame(f0,bg='lightgrey')
    f1.place(x=75,y=60,relheight=0.65,relwidth=0.28)
    
    
    lb1=Label(f1,text='Product Name',font=def_font,bg='lightgrey')
    lb1.place(x=0,y=0)
    entry1=Entry(f1,font=('times new roman',13),relief=FLAT,textvariable=search_name,width=20)
    entry1.place(x=2,y=25,relwidth=0.75)
    search_name.trace('w',callback)
    search_btn=Button(f1,text='Search',command=item_search,relief=FLAT,font=('times new roman',11),bd=0,bg='#7f95d8',width=8)
    search_btn.place(y=25,relx=0.78)

    list_box=Listbox(f1,selectmode=EXTENDED,font=def_font,relief=FLAT)
    list_box.config(activestyle='none')
    x1=Label(list_box,text='*** Enter the Item Name Above ***',font=('times new roman',11),bg='white')
    x1.place(relx=0.11,rely=0.45,relwidth=0.8)
    list_box.config(activestyle='none',selectbackground='grey',bd=0)
    list_box.place(x=0,y=80,relwidth=1,relheight=0.6)
    list_box.bind('<Return>',get_item)
    entry1.bind('<Return>',item_search)
    entry1.focus_set()

    lb2=Label(f1,text='Selected Product',font=def_font,bg='lightgrey')
    lb2.place(x=0,rely=0.8)
    lb21=Label(f1,text='Discount',font=def_font,bg='lightgrey')
    lb21.place(relx=0.55,rely=0.8)
    entry21=Entry(f1,font=('times new roman',13),relief=FLAT,textvariable=discount,width=20)
    entry21.place(relx=0.55,rely=0.85,relwidth=0.22)
    entry21.bind('<Return>',get_quantity)
    c1=Canvas(f1,height=20)
    c1.place(x=2,rely=0.85,relwidth=0.52)
    lb3=Label(f1,text='Count',font=def_font,bg='lightgrey')
    lb3.place(relx=0.8,rely=0.8)
    entry2=Entry(f1,font=('times new roman',13),relief=FLAT,textvariable=qty,width=20)
    entry2.place(relx=0.8,rely=0.85,relwidth=0.2)
    entry2.bind('<Return>',get_quantity)
    
    f3=Frame(f1,bg='lightgrey')
    add_btn=Button(f3,text='Add',font=('times new roman',11),command=get_quantity,width=8,bd=0,relief=FLAT,bg='#7f95d8')
    add_btn.pack(side=LEFT,padx=2)
    cancel_btn=Button(f3,text='Cancel',font=('times new roman',11),command=cancel,bd=0,width=8,relief=FLAT,bg='#7f95d8')
    cancel_btn.pack(pady=3,side=RIGHT)
    f3.pack(side=BOTTOM,anchor=SE)

    
    f2=Frame(f0,bg='lightgrey')
    f2.place(relx=0.62,rely=0.08,relheight=0.63,relwidth=0.25)

    lb3=Label(f2,text='Bill',font=def_font,bg='#9fabcf')
    lb3.place(relx=0,y=45,relwidth=1)
    bill_list=Listbox(f2,relief=FLAT,selectmode=EXTENDED,font='arial 13')
    bill_list.config(activestyle='none',selectbackground='grey',bd=0)
    bill_list.place(x=0,y=70,relwidth=1,relheight=0.75)
    
    f4=Frame(f2,bg='lightgrey')
    print_btn=Button(f4,text="Print",font=('times new roman',11),command=get_money,width=8,bd=0,relief=FLAT,bg='#7f95d8')
    print_btn.pack(side=LEFT)
    remove_btn=Button(f4,text='Remove',font=('times new roman',11),command=remove,width=8,bd=0,relief=FLAT,bg='#7f95d8')
    remove_btn.pack(side=LEFT,padx=2)
    clear_btn=Button(f4,text='Clear',font=('times new roman',11),command=clear,width=8,bd=0,relief=FLAT,bg='#7f95d8')
    clear_btn.pack(side=RIGHT)
    f4.pack(side=BOTTOM,anchor=SE)
    
    f5=Frame(f0,bg='lightgrey')
    f5.place(relx=0.385,rely=0.75,relheight=0.1,relwidth=0.5)

    lb5=Label(f5,text='Amount',font=def_font,bg='lightgrey')
    lb5.place(relx=0,rely=0)
    c2=Canvas(f5,bd=0,height=20)
    c2.place(relx=0,rely=0.3,relwidth=0.32)

    lb6=Label(f5,text='Paid',font=def_font,bg='lightgrey')
    lb6.place(relx=0.33,rely=0)
    entry3=Entry(f5,font=('times new roman',13),relief=FLAT,textvariable=paid,width=20)
    entry3.place(relx=0.33,rely=0.3,relwidth=0.32)
    entry3.bind('<Return>',get_balance)

    lb7=Label(f5,text='Balance',font=def_font,bg='lightgrey')
    lb7.place(relx=0.66,rely=0)
    c3=Canvas(f5,bd=0,height=20)
    c3.place(relx=0.66,rely=0.3,relwidth=0.34)

    remove_btn.bind('<Enter>',hover)
    remove_btn.bind('<Leave>',cancel_hover)
    cancel_btn.bind('<Enter>',hover)
    cancel_btn.bind('<Leave>',cancel_hover)
    clear_btn.bind('<Enter>',hover)
    clear_btn.bind('<Leave>',cancel_hover)
    add_btn.bind('<Enter>',hover)
    add_btn.bind('<Leave>',cancel_hover)
    search_btn.bind('<Enter>',hover)
    search_btn.bind('<Leave>',cancel_hover)
    print_btn.bind('<Enter>',hover)
    print_btn.bind('<Leave>',cancel_hover)

def update_cost(e=None):
    global act_list,item_name,up_cost,entry1,lb8
    q=int(up_cost.get())
    
    for i in range(2,work_sheet.max_row+1):
        if item_name==work_sheet.cell(row=i,column=2).value:
            work_sheet.cell(row=i,column=3).value=q
            book.save('item_details.xlsx')
            break
    
    act_list.insert(END,item_name+'`s cost is updated to Rs,'+str(work_sheet.cell(row=i,column=3).value))
    entry1.focus_set()

def update_count(Event=None):
    global act_list,item_name,qty,entry1,lb8
    q=int(qty.get())
    
    for i in range(2,work_sheet.max_row+1):
        if item_name==work_sheet.cell(row=i,column=2).value:
            work_sheet.cell(row=i,column=4).value+=q
            book.save('D:\python\project one\item_details.xlsx')
            break
    
    act_list.insert(END,' > '+item_name+'`s count is updated to '+str(work_sheet.cell(row=i,column=4).value))
    lb8=Label(f1,text='Total number of updated items :')
    entry1.focus_set()



def stock():
    global search_name,f0,f1,f2,f6,l8,flag2,flag1,logo,lb1,entry1,search_btn,list_box,x1,c1,lb2,lb3,lb21,entry2,entry21,update_btn,cancel_btn,act_list,update_btn

    if flag2==True:
        flag2=False
        flag1=True
    else:
        return
    f0.destroy()
    search_name.set('')
    qty.set('')
    f0=Frame(app,bg='lightgrey')
    f0.place(x=87,y=81,relheight=1,relwidth=1)
    logo=Label(f0,image=pic,bg='lightgrey')
    logo.place(x=-100,y=-100,relwidth=1,relheight=1)
    Label(f0,text='stock Update',font=def_font,bg='#9fabcf').place(x=-50,relwidth=1)
    l8=Label(f0,text='Careful before click a button...!!!',font=def_font,bg='lightgrey')
    l8.place(rely=0.83,relwidth=1)
    f1=Frame(f0,bg='lightgrey')
    f1.place(x=75,y=60,relheight=0.65,relwidth=0.28)
    
    lb1=Label(f1,text='Product Name',font=def_font,bg='lightgrey')
    lb1.place(x=0,y=0)
    entry1=Entry(f1,font=('times new roman',13),relief=FLAT,textvariable=search_name,width=20)
    entry1.place(x=2,y=25,relwidth=0.7)
    search_btn=Button(f1,text='Search',command=search_item,relief=FLAT,font=('times new roman',11),bd=0,bg='#7f95d8',width=8)
    search_btn.place(y=25,relx=0.76)

    list_box=Listbox(f1,selectmode=EXTENDED,font=def_font,relief=FLAT)
    list_box.config(activestyle='none')
    x1=Label(list_box,text='*** Enter the Item Name Above ***',font=('times new roman',11),bg='white')
    x1.place(relx=0.11,rely=0.45,relwidth=0.8)
    list_box.config(activestyle='none',selectbackground='grey',bd=0)
    list_box.place(x=0,y=80,relwidth=1,relheight=0.6)
    list_box.bind('<Return>',get_item)
    entry1.bind('<Return>',search_item)
    entry1.focus_set()

    lb2=Label(f1,text='Selected Product',font=def_font,bg='lightgrey')
    lb2.place(x=0,rely=0.8)
    c1=Canvas(f1,height=20)
    c1.place(x=2,rely=0.85,relwidth=0.5)
    lb3=Label(f1,text='Count',font=def_font,bg='lightgrey')
    lb3.place(relx=0.8,rely=0.8)
    lb21=Label(f1,text='Cost',font=def_font,bg='lightgrey')
    lb21.place(relx=0.55,rely=0.8)
    entry21=Entry(f1,font=('times new roman',13),relief=FLAT,textvariable=up_cost,width=20)
    entry21.place(relx=0.55,rely=0.85,relwidth=0.2)
    entry21.bind('<Return>',update_cost)
    entry2=Entry(f1,font=('times new roman',13),relief=FLAT,textvariable=qty,width=20)
    entry2.place(relx=0.8,rely=0.85,relwidth=0.2)
    entry2.bind('<Return>',update_count)
    
    f3=Frame(f1,bg='lightgrey')
    update_btn=Button(f3,text='Update',font=('times new roman',11),command=update_count,width=8,bd=0,relief=FLAT,bg='#7f95d8')
    update_btn.pack(side=LEFT,padx=2)
    cancel_btn=Button(f3,text='Cancel',font=('times new roman',11),command=cancel,bd=0,width=8,relief=FLAT,bg='#7f95d8')
    cancel_btn.pack(pady=3,side=RIGHT)
    f3.pack(side=BOTTOM,anchor=SE)

    f2=Frame(f0,bg='lightgrey')
    f2.place(relx=0.62,rely=0.12,relheight=0.63,relwidth=0.25)

    lb3=Label(f2,text='Activities',font=def_font,bg='#9fabcf')
    lb3.place(relx=0,y=0,relwidth=1)
    act_list=Listbox(f2,relief=RIDGE,selectmode=EXTENDED,font='timesnewroman 10',bg='#aacfe2')
    act_list.config(activestyle='none',selectbackground='grey',bd=0)
    act_list.place(x=0,y=25,relwidth=1,relheight=1)


    update_btn.bind('<Enter>',hover)
    update_btn.bind('<Leave>',cancel_hover)
    cancel_btn.bind('<Enter>',hover)
    cancel_btn.bind('<Leave>',cancel_hover)
    search_btn.bind('<Enter>',hover)
    search_btn.bind('<Leave>',cancel_hover)




img=Image.open("logo.png")
pic=ImageTk.PhotoImage(img)
img1=Image.open("1.png")
l_pic=ImageTk.PhotoImage(img1)
logo=Label(app,image=pic,bg='lightgrey')
logo.place(x=-30,y=0,relwidth=1,relheight=1)
l2=Frame(app,bg='#4d4d4d',height=80)
l2.pack(side=TOP,fill='x')
l1=Frame(app,bg='#676767')
l1.pack(side=LEFT,fill='y')
Label(l2,image=l_pic,bg='#4d4d4d').place(x=0,y=0)

f0=Frame(app)
f1=Frame(f0)
f2=Frame(f0)
f3=Frame(f1)
f4=Frame(f2)
f5=Frame(f0)

x1=Label(f1)
lb1=Label(f1)
lb2=Label(f1)
lb3=Label(f1)
lb21=Label(f1)
lb4=Label(f1)
lb5=Label(f0)
lb6=Label(f0)
lb7=Label(f0)
lb8=Label(f0)
l8=Label(f0)
l9=Label()

c1=Canvas(f1)
c2=Canvas(f5)
c3=Canvas(f5)
c7=Canvas(l8)

entry1=Entry(f1)
entry2=Entry(f1)
entry3=Entry(f5)
entry21=Entry(f1)

bill_btn=Button(l1,text='Bill',relief=FLAT,command=billing,height=3,bd=0,width=8,font=def_font,bg='#939393')
bill_btn.pack(fill='x',pady=1)
bill_btn.bind('<Enter>',hover)
bill_btn.bind('<Leave>',cancel_hover)
stock_btn=Button(l1,text='stock',relief=FLAT,command=stock,height=3,bd=0,font=def_font,bg='#939393')
stock_btn.pack(fill='x',pady=1)
stock_btn.bind('<Enter>',hover)
stock_btn.bind('<Leave>',cancel_hover)

list_box=Listbox(f1)
bill_list=Listbox(f2)
act_list=Listbox(f2)
x1=Label(list_box)

back_btn=Button(app)
add_btn=Button(f3)
cancel_btn=Button(f3)
search_btn=Button(f1)
remove_btn=Button(f4)
clear_btn=Button(f4)
print_btn=Button(f4)
update_btn=Button(f3)

search_name=StringVar()
qty=StringVar()
qty=StringVar()
income=IntVar()
qty.set('')
up_cost=StringVar()
discount=IntVar()
up_cost.set('')
bill=[]
bill_sno=0
out=''
item_name=''
total_amount=0
paid=StringVar()
flag1=True
flag2=True

app.mainloop()
